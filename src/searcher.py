import os
import openpyxl
import telebot
import urllib3
from urllib3.exceptions import InsecureRequestWarning

# Отключение предупреждений о проверке SSL
urllib3.disable_warnings(InsecureRequestWarning)

# Конфигурация
TOKEN = "8164754969:AAGOgO3oEccKdxbreCv9TpRiSw3Nby5qv04"
BASE_FILE_PATH = "files/base.xlsx"
PROXY_URL = None

# Создаем папку files, если ее нет
if not os.path.exists("files"):
    os.makedirs("files")


def convert_xls_to_xlsx_if_needed():
    """Проверяет формат файла и при необходимости конвертирует"""
    if not os.path.exists(BASE_FILE_PATH):
        xls_path = BASE_FILE_PATH.replace('.xlsx', '.xls')
        if os.path.exists(xls_path):
            try:
                import xlrd
                from openpyxl import Workbook

                wb_xls = xlrd.open_workbook(xls_path)
                sheet = wb_xls.sheet_by_index(0)
                wb_xlsx = Workbook()
                ws = wb_xlsx.active

                for row in range(sheet.nrows):
                    for col in range(sheet.ncols):
                        ws.cell(row=row + 1, column=col + 1).value = sheet.cell_value(row, col)

                wb_xlsx.save(BASE_FILE_PATH)
                print(f"Файл {xls_path} успешно конвертирован в {BASE_FILE_PATH}")
                return True
            except Exception as e:
                print(f"Ошибка при конвертации файла: {e}")
                return False
    return False


def clean_article_column():
    """Удаляет пробелы только во втором столбце файла Excel (артикулы)"""
    try:
        if not os.path.exists(BASE_FILE_PATH):
            if not convert_xls_to_xlsx_if_needed():
                raise FileNotFoundError(f"Файл {BASE_FILE_PATH} не найден")

        workbook = openpyxl.load_workbook(BASE_FILE_PATH)
        sheet = workbook.active

        # Очищаем только второй столбец (артикулы)
        for row in sheet.iter_rows(min_col=2, max_col=2):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell.value = cell.value.replace(" ", "")

        workbook.save(BASE_FILE_PATH)
        print("Пробелы во втором столбце (артикулы) успешно удалены.")
    except Exception as e:
        print(f"Произошла ошибка при обработке файла: {str(e)}")
        raise


# Инициализация бота
bot = telebot.TeleBot(TOKEN, skip_pending=True, threaded=False)


def format_price(price):
    """Форматирует цену с разделителями тысяч"""
    try:
        return "{:,.2f}".format(float(price)).replace(",", "").replace(".", ",")
    except:
        return str(price)


def normalize_article(article):
    """Нормализует артикул: удаляет пробелы и приводит к нижнему регистру"""
    return article.replace(" ", "").strip().lower()


def search_in_base(article: str) -> dict:
    """Поиск артикула в базе и возврат информации о товаре"""
    try:
        if not os.path.exists(BASE_FILE_PATH):
            if not convert_xls_to_xlsx_if_needed():
                return {"error": "Файл базы данных не найден"}

        workbook = openpyxl.load_workbook(BASE_FILE_PATH)
        sheet = workbook.active
        article = normalize_article(article)
        results = []

        for row in sheet.iter_rows(values_only=True):
            # Нормализуем только второй столбец (артикулы)
            col2 = normalize_article(str(row[1])) if row[1] is not None else ""
            # Третий столбец (наименование) оставляем как есть, без нормализации
            col3 = str(row[2]) if row[2] is not None else ""

            # Ищем вхождение в нормализованном артикуле (col2) или в оригинальном наименовании (col3)
            if article in col2 or article in col3.lower().replace(" ", ""):
                item = {
                    "name": col3,  # Оригинальное наименование с пробелами
                    "price": format_price(row[3]) if len(row) > 3 and row[3] is not None else "Не указана",
                    "quantity": str(row[4]) if len(row) > 4 and row[4] is not None else "Не указано",
                    "article": str(row[1]) if row[1] is not None else "Не указан"
                }
                results.append(item)

        if not results:
            return {"error": "Артикул не найден"}

        return {"results": results}

    except Exception as e:
        return {"error": f"Ошибка: {str(e)}"}


# Обработчик текстовых сообщений
@bot.message_handler(func=lambda message: True)
def handle_message(message):
    """Обработчик текстовых сообщений"""
    article = message.text.strip()
    search_result = search_in_base(article)

    if "error" in search_result:
        bot.send_message(message.chat.id, search_result["error"])
    else:
        for item in search_result["results"]:
            response = (
                f"📌 Наименование: {item['name']}\n"
                f"💰 Цена за штуку: {item['price']} руб.\n"
                f"📦 Количество на складе: {item['quantity']}"
            )
            bot.send_message(message.chat.id, response)


if __name__ == "__main__":
    try:
        print("Запуск бота...")
        clean_article_column()
        print("База данных готова к поиску")
        bot.polling(none_stop=True, skip_pending=True)
        print("Бот успешно запущен!")
    except KeyboardInterrupt:
        print("Бот остановлен пользователем")
    except Exception as e:
        print(f"Критическая ошибка: {e}")